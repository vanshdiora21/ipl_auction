"""
create_template.py
==================
Run this ONCE to create your players.xlsx with the correct format.
Then edit the player names/teams directly in Excel.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Players"

    # ── Styling helpers ──────────────────────────────────
    header_fill  = PatternFill("solid", start_color="1F4E79")
    header_font  = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    center       = Alignment(horizontal="center", vertical="center")
    left         = Alignment(horizontal="left",   vertical="center")
    thin         = Side(style="thin", color="BFBFBF")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Headers ──────────────────────────────────────────
    headers = ["Player Name", "Team", "← Date columns fill in automatically →"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = center if col > 1 else left
        cell.border = border

    # ── Sample players (replace with yours) ─────────────
    sample_players = [
        ("Virat Kohli",       "India"),
        ("Rohit Sharma",      "India"),
        ("Jasprit Bumrah",    "India"),
        ("MS Dhoni",          "India"),
        ("KL Rahul",          "India"),
        ("Shubman Gill",      "India"),
        ("Hardik Pandya",     "India"),
        ("Ravindra Jadeja",   "India"),
        ("Mohammed Shami",    "India"),
        ("Shreyas Iyer",      "India"),
        ("Jos Buttler",       "England"),
        ("Ben Stokes",        "England"),
        ("Babar Azam",        "Pakistan"),
        ("Steve Smith",       "Australia"),
        ("Pat Cummins",       "Australia"),
    ]

    name_fill    = PatternFill("solid", start_color="DEEAF1")   # light blue
    alt_fill     = PatternFill("solid", start_color="F5FAFD")   # very light blue
    name_font    = Font(name="Arial", size=10, color="1F4E79", bold=True)
    team_font    = Font(name="Arial", size=10, color="404040")

    for idx, (name, team) in enumerate(sample_players, start=2):
        fill = name_fill if idx % 2 == 0 else alt_fill

        nc = ws.cell(row=idx, column=1, value=name)
        nc.font      = name_font
        nc.fill      = fill
        nc.alignment = left
        nc.border    = border

        tc = ws.cell(row=idx, column=2, value=team)
        tc.font      = team_font
        tc.fill      = fill
        tc.alignment = center
        tc.border    = border

    # ── Column widths ────────────────────────────────────
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 36
    ws.row_dimensions[1].height     = 22

    # ── Freeze top row ───────────────────────────────────
    ws.freeze_panes = "A2"

    # ── Instructions sheet ───────────────────────────────
    info = wb.create_sheet("Instructions")
    info["A1"] = "CREX Fantasy Points Tracker — Instructions"
    info["A1"].font = Font(bold=True, size=14, name="Arial", color="1F4E79")

    steps = [
        ("Step 1", "Edit the 'Players' sheet — add/remove players in Column A, team in Column B."),
        ("Step 2", "Make sure Python 3 is installed on your computer."),
        ("Step 3", "Install dependencies once:  pip install selenium openpyxl webdriver-manager beautifulsoup4"),
        ("Step 4", "Every day, run:  python fantasy_points_tracker.py"),
        ("Step 5", "A new date column will appear automatically with today's points!"),
        ("Note",   "Points show as green if found, grey if no match today, red if an error occurred."),
        ("Tip",    "Keep this file in the same folder as fantasy_points_tracker.py."),
    ]

    for r, (label, text) in enumerate(steps, start=3):
        lc = info.cell(row=r, column=1, value=label)
        lc.font = Font(bold=True, name="Arial", size=10, color="1F4E79")
        lc.alignment = Alignment(vertical="center")

        tc = info.cell(row=r, column=2, value=text)
        tc.font = Font(name="Arial", size=10)
        tc.alignment = Alignment(wrap_text=True, vertical="center")
        info.row_dimensions[r].height = 20

    info.column_dimensions["A"].width = 12
    info.column_dimensions["B"].width = 80

    # ── Save ─────────────────────────────────────────────
    wb.save("players.xlsx")
    print("✅ Created players.xlsx")
    print("   → Open it, edit your player list, then run: python fantasy_points_tracker.py")

if __name__ == "__main__":
    create_template()
