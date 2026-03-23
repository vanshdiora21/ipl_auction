"""
CREX Fantasy Points Tracker — Final Version
=============================================
Fetches today's IPL 2026 scorecard from crex.com and calculates
fantasy points using the standard Dream11/CREX T20 point system.
Logs results to players.xlsx under today's date column.

HOW TO RUN LOCALLY:
    source venv/bin/activate
    python3 fantasy_points_tracker.py

REQUIREMENTS:
    pip install requests openpyxl beautifulsoup4
"""

import re
import sys
import os
from datetime import datetime, date
from pathlib import Path

try:
    import requests
    from bs4 import BeautifulSoup
except ImportError:
    sys.exit("Run: pip install requests beautifulsoup4")

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    sys.exit("Run: pip install openpyxl")

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────────────────────────────────────
EXCEL_FILE     = "players.xlsx"
DATA_START_ROW = 2

# IPL 2026 full schedule — date → list of CREX scorecard paths
# Dates are in IST. Update this dict as more fixtures are announced.
IPL_SCHEDULE = {
    "2026-03-28": ["scoreboard/10XZ/1PW/1st-Match/K/L/rcb-vs-srh-1st-match-indian-premier-league-2026"],
    "2026-03-29": ["scoreboard/10Y0/1PW/2nd-Match/F/J/kkr-vs-mi-2nd-match-indian-premier-league-2026"],
    "2026-03-30": ["scoreboard/10Y1/1PW/3rd-Match/G/M/csk-vs-rr-3rd-match-indian-premier-league-2026"],
    "2026-03-31": ["scoreboard/10Y2/1PW/4th-Match/I/KB/gt-vs-pbks-4th-match-indian-premier-league-2026"],
    "2026-04-01": ["scoreboard/10Y3/1PW/5th-Match/H/KC/dc-vs-lsg-5th-match-indian-premier-league-2026"],
    "2026-04-02": ["scoreboard/10Y4/1PW/6th-Match/J/L/kkr-vs-srh-6th-match-indian-premier-league-2026"],
    "2026-04-03": ["scoreboard/10Y5/1PW/7th-Match/G/I/csk-vs-pbks-7th-match-indian-premier-league-2026"],
    "2026-04-04": [
        "scoreboard/10Y6/1PW/8th-Match/F/H/dc-vs-mi-8th-match-indian-premier-league-2026",
        "scoreboard/10Y7/1PW/9th-Match/KB/M/gt-vs-rr-9th-match-indian-premier-league-2026",
    ],
    "2026-04-05": [
        "scoreboard/10Y8/1PW/10th-Match/KC/L/lsg-vs-srh-10th-match-indian-premier-league-2026",
        "scoreboard/10Y9/1PW/11th-Match/G/K/csk-vs-rcb-11th-match-indian-premier-league-2026",
    ],
    "2026-04-06": ["scoreboard/10YA/1PW/12th-Match/I/J/kkr-vs-pbks-12th-match-indian-premier-league-2026"],
    "2026-04-07": ["scoreboard/10YB/1PW/13th-Match/F/M/mi-vs-rr-13th-match-indian-premier-league-2026"],
    "2026-04-08": ["scoreboard/10YC/1PW/14th-Match/H/KB/dc-vs-gt-14th-match-indian-premier-league-2026"],
    "2026-04-09": ["scoreboard/10YD/1PW/15th-Match/J/KC/kkr-vs-lsg-15th-match-indian-premier-league-2026"],
    "2026-04-10": ["scoreboard/10YE/1PW/16th-Match/K/M/rcb-vs-rr-16th-match-indian-premier-league-2026"],
    "2026-04-11": [
        "scoreboard/10YF/1PW/17th-Match/I/L/pbks-vs-srh-17th-match-indian-premier-league-2026",
        "scoreboard/10YG/1PW/18th-Match/G/H/csk-vs-dc-18th-match-indian-premier-league-2026",
    ],
    "2026-04-12": [
        "scoreboard/10YH/1PW/19th-Match/KB/KC/gt-vs-lsg-19th-match-indian-premier-league-2026",
        "scoreboard/10YI/1PW/20th-Match/F/K/mi-vs-rcb-20th-match-indian-premier-league-2026",
    ],
}

CREX_BASE = "https://crex.com"
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Referer": "https://crex.com/",
}

# ─────────────────────────────────────────────────────────────────────────────
# FANTASY POINT SYSTEM  (Standard Dream11 / CREX T20 rules)
# ─────────────────────────────────────────────────────────────────────────────
PTS_PLAYING        =  4
PTS_RUN            =  1
PTS_BOUNDARY       =  1
PTS_SIX            =  2
PTS_FIFTY          =  8
PTS_CENTURY        = 16
PTS_DUCK           = -2
SR_BONUS_170       =  6
SR_BONUS_150       =  4
SR_BONUS_130       =  2
SR_PENALTY_60      = -2
SR_PENALTY_50      = -4
SR_PENALTY_40      = -6
PTS_WICKET         = 25
PTS_LBW_BOWLED     =  8
PTS_3WICKETS       =  4
PTS_4WICKETS       =  8
PTS_5WICKETS       = 16
PTS_MAIDEN         = 12
ER_BONUS_4         =  6
ER_BONUS_5         =  4
ER_BONUS_7         =  2
ER_PENALTY_12      = -2
ER_PENALTY_13      = -4
ER_PENALTY_14      = -6
PTS_CATCH          =  8
PTS_STUMPING       = 12
PTS_RUN_OUT_DIRECT = 12


def calc_batting_pts(runs, balls, fours, sixes, dismissed, role):
    if balls == 0 and runs == 0:
        return 0
    pts  = runs * PTS_RUN
    pts += fours * PTS_BOUNDARY
    pts += sixes * PTS_SIX
    if runs >= 100:
        pts += PTS_CENTURY
    elif runs >= 50:
        pts += PTS_FIFTY
    if dismissed and runs == 0 and role.lower() != "bowler":
        pts += PTS_DUCK
    if balls >= 10:
        sr = (runs / balls) * 100
        if   sr >= 170: pts += SR_BONUS_170
        elif sr >= 150: pts += SR_BONUS_150
        elif sr >= 130: pts += SR_BONUS_130
        elif sr <   40: pts += SR_PENALTY_40
        elif sr <   50: pts += SR_PENALTY_50
        elif sr <   60: pts += SR_PENALTY_60
    return pts


def calc_bowling_pts(wickets, lbw_bowled, overs, maidens, runs_given):
    if overs == 0:
        return 0
    pts  = wickets * PTS_WICKET
    pts += lbw_bowled * PTS_LBW_BOWLED
    pts += maidens * PTS_MAIDEN
    if   wickets >= 5: pts += PTS_5WICKETS
    elif wickets >= 4: pts += PTS_4WICKETS
    elif wickets >= 3: pts += PTS_3WICKETS
    if overs >= 2:
        er = runs_given / overs
        if   er <  4: pts += ER_BONUS_4
        elif er <  5: pts += ER_BONUS_5
        elif er <  7: pts += ER_BONUS_7
        elif er > 14: pts += ER_PENALTY_14
        elif er > 13: pts += ER_PENALTY_13
        elif er > 12: pts += ER_PENALTY_12
    return pts


# ─────────────────────────────────────────────────────────────────────────────
# SCORECARD SCRAPING
# ─────────────────────────────────────────────────────────────────────────────

def _blank():
    return {
        "runs": 0, "balls": 0, "fours": 0, "sixes": 0,
        "dismissed": False, "batted": False,
        "wickets": 0, "overs": 0.0, "maidens": 0,
        "runs_given": 0, "lbw_bowled": 0, "bowled": False,
        "catches": 0, "stumpings": 0, "run_outs": 0,
        "role": "batsman",
    }


def fetch_scorecard(path: str) -> dict:
    url = f"{CREX_BASE}/{path}/scorecard"
    print(f"  Fetching: {url}")
    try:
        r = requests.get(url, headers=HEADERS, timeout=15)
        r.raise_for_status()
    except requests.exceptions.RequestException as e:
        print(f"  ⚠ Could not fetch scorecard: {e}")
        return {}

    soup = BeautifulSoup(r.text, "html.parser")
    players = {}

    # Parse batting rows: Name | How Out | R | B | 4s | 6s | SR
    for row in soup.select("table tr, .batting-row, [class*='bat']"):
        cells = [td.get_text(strip=True) for td in row.find_all(["td", "th"])]
        if len(cells) < 6:
            continue
        try:
            runs  = int(cells[2])
            balls = int(cells[3])
            fours = int(cells[4])
            sixes = int(cells[5])
        except (ValueError, IndexError):
            continue
        name    = cells[0].strip()
        how_out = cells[1].strip().lower() if len(cells) > 1 else ""
        dismissed = how_out not in ("not out", "dnb", "")
        if name and len(name) > 3 and not name.lower().startswith("extra"):
            if name not in players:
                players[name] = _blank()
            players[name].update({
                "runs": runs, "balls": balls,
                "fours": fours, "sixes": sixes,
                "dismissed": dismissed, "batted": True,
            })

    # Parse bowling rows: Name | O | M | R | W | Econ
    for row in soup.select("table tr, .bowling-row, [class*='bowl']"):
        cells = [td.get_text(strip=True) for td in row.find_all(["td", "th"])]
        if len(cells) < 5:
            continue
        try:
            overs   = float(cells[1])
            maidens = int(cells[2])
            runs_g  = int(cells[3])
            wickets = int(cells[4])
        except (ValueError, IndexError):
            continue
        name = cells[0].strip()
        if name and len(name) > 3:
            if name not in players:
                players[name] = _blank()
            players[name].update({
                "overs": overs, "maidens": maidens,
                "runs_given": runs_g, "wickets": wickets,
                "bowled": True,
            })

    return players


def compute_fantasy_points(players: dict) -> dict:
    result = {}
    for name, p in players.items():
        pts  = PTS_PLAYING
        pts += calc_batting_pts(
            p["runs"], p["balls"], p["fours"], p["sixes"],
            p["dismissed"], p.get("role", "batsman")
        )
        pts += calc_bowling_pts(
            p["wickets"], p.get("lbw_bowled", 0),
            p["overs"], p["maidens"], p["runs_given"]
        )
        pts += p["catches"]   * PTS_CATCH
        pts += p["stumpings"] * PTS_STUMPING
        pts += p["run_outs"]  * PTS_RUN_OUT_DIRECT
        result[name] = round(pts, 1)
    return result


# ─────────────────────────────────────────────────────────────────────────────
# PLAYER MATCHING
# ─────────────────────────────────────────────────────────────────────────────

def normalize(name: str) -> str:
    return re.sub(r'\s+', ' ', re.sub(r'[^a-z\s]', '', name.lower())).strip()


def match_players(your_players, scraped_points) -> dict:
    scraped_norm = {normalize(k): v for k, v in scraped_points.items()}
    results = {}
    for player in your_players:
        name  = player["name"]
        norm  = normalize(name)
        parts = norm.split()
        pts   = None
        # 1. Exact match
        if norm in scraped_norm:
            pts = scraped_norm[norm]
        else:
            # 2. Last name
            for sn, sv in scraped_norm.items():
                if parts and parts[-1] in sn.split():
                    pts = sv; break
            # 3. First name
            if pts is None and len(parts) > 1:
                for sn, sv in scraped_norm.items():
                    snp = sn.split()
                    if snp and snp[0] == parts[0]:
                        pts = sv; break
        results[name] = str(pts) if pts is not None else "No match today"
    return results


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL I/O
# ─────────────────────────────────────────────────────────────────────────────

def load_players(wb):
    ws = wb["Players"]
    players = []
    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        if not row[0]:
            continue
        players.append({
            "name": str(row[0]).strip(),
            "team": str(row[1]).strip() if len(row) > 1 and row[1] else "",
        })
    return players


def get_or_create_date_col(ws, today_str):
    for col in range(3, ws.max_column + 2):
        cell = ws.cell(row=1, column=col)
        if cell.value is None:
            cell.value = today_str
            cell.font      = Font(bold=True, color="FFFFFF", name="Arial")
            cell.fill      = PatternFill("solid", start_color="1F4E79")
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[cell.column_letter].width = 14
            return col
        if str(cell.value) == today_str:
            return col
    return ws.max_column + 1


def write_results(ws, players, results, date_col):
    for row_idx, player in enumerate(players, start=DATA_START_ROW):
        val  = results.get(player["name"], "N/A")
        cell = ws.cell(row=row_idx, column=date_col, value=val)
        cell.alignment = Alignment(horizontal="center")
        cell.font      = Font(name="Arial", size=10)
        try:
            float(val)
            cell.fill = PatternFill("solid", start_color="E2EFDA")
            cell.font = Font(color="375623", name="Arial", size=10, bold=True)
        except ValueError:
            cell.fill = PatternFill("solid", start_color="F2F2F2")


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    # GitHub Actions runs in UTC; IST = UTC+5:30
    # We use IST date to match the schedule correctly
    from datetime import timezone, timedelta
    IST = timezone(timedelta(hours=5, minutes=30))
    now_ist   = datetime.now(IST)
    today_key = now_ist.strftime("%Y-%m-%d")
    today_str = now_ist.strftime("%d-%b-%Y")

    print(f"\n🏏  CREX Fantasy Points Tracker  —  {today_str} (IST)")
    print("=" * 52)

    excel_path = Path(EXCEL_FILE)
    if not excel_path.exists():
        sys.exit(f"❌  '{EXCEL_FILE}' not found.")

    wb = load_workbook(excel_path)
    if "Players" not in wb.sheetnames:
        sys.exit("❌  No sheet named 'Players' found.")

    players = load_players(wb)
    print(f"📋  {len(players)} players loaded.\n")

    match_paths = IPL_SCHEDULE.get(today_key, [])
    if not match_paths:
        print(f"📅  No IPL match scheduled today ({today_str}). Saving 'No match today'.")
        results = {p["name"]: "No match today" for p in players}
    else:
        print(f"📅  {len(match_paths)} match(es) today. Fetching scorecards...\n")
        all_raw = {}
        for path in match_paths:
            raw = fetch_scorecard(path)
            print(f"     → {len(raw)} players parsed.")
            all_raw.update(raw)

        if not all_raw:
            print("\n  ⚠ Scorecard not yet available (match may not have finished).")
            results = {p["name"]: "Scorecard unavailable" for p in players}
        else:
            print(f"\n📊  Computing fantasy points...")
            fantasy_pts = compute_fantasy_points(all_raw)

            print("\n  All scraped players:")
            for pname, pts in sorted(fantasy_pts.items(), key=lambda x: -x[1]):
                print(f"    {pname}: {pts}")

            print(f"\n🔗  Matching to your {len(players)} players...")
            results = match_players(players, fantasy_pts)

    found = 0
    print()
    for name, pts in results.items():
        has  = pts not in ("N/A", "No match today", "Scorecard unavailable")
        found += has
        print(f"  {'✅' if has else '—'}  {name}: {pts}")

    ws = wb["Players"]
    date_col = get_or_create_date_col(ws, today_str)
    write_results(ws, players, results, date_col)
    wb.save(excel_path)

    print(f"\n✅  Done! {found}/{len(players)} players matched.")
    print(f"💾  Saved → '{EXCEL_FILE}'  column '{today_str}'\n")


if __name__ == "__main__":
    main()
    